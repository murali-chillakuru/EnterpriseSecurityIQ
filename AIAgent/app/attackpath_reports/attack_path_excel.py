"""
Attack Path Detection — Excel Export.

Generates a multi-sheet Excel workbook:
  1. Summary — overview stats and posture
  2. Attack Paths — all paths with full details
  3. MITRE Coverage — technique/tactic mapping
  4. Remediation Tracker — actionable remediation with status/owner/due-date columns
  5. Compliance Coverage — frameworks mapped to findings
  6. Evidence Summary — data sources and collection metadata
"""
from __future__ import annotations

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

_SEV_FILLS = {
    "critical": "FFD13438",
    "high": "FFFF8C00",
    "medium": "FFFFC107",
    "low": "FF0078D4",
    "informational": "FF6B6B6B",
}

_FRAMEWORKS = ("NIST-800-53", "CIS", "HIPAA", "PCI-DSS", "ISO-27001", "SOC-2")


def _apply_header(cell, header_font, header_fill):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, col_count, max_rows=80):
    for c in range(1, col_count + 1):
        letter = get_column_letter(c)
        w = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, min(max_rows, ws.max_row + 1))),
            default=10,
        )
        ws.column_dimensions[letter].width = min(w + 4, 65)


def generate_excel_report(assessment: dict, output_path: str) -> None:
    """Generate Excel report from assessment dict."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF0078D4", end_color="FF0078D4", fill_type="solid")
    accent_fill = PatternFill(start_color="FF107C10", end_color="FF107C10", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

    summary = assessment.get("Summary", {})
    paths = assessment.get("Paths", [])
    evidence = assessment.get("Evidence", {})

    # ── Sheet 1: Summary ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "0078D4"

    rows = [
        ("Metric", "Value"),
        ("Tenant ID", assessment.get("TenantId", "")),
        ("Assessment Timestamp", assessment.get("AssessmentTimestamp", "")),
        ("Total Attack Paths", summary.get("TotalPaths", 0)),
        ("Overall Risk Score", summary.get("OverallRiskScore", 0)),
        ("Overall Severity", summary.get("OverallSeverity", "").upper()),
        ("Critical Paths", summary.get("SeverityCounts", {}).get("critical", 0)),
        ("High Paths", summary.get("SeverityCounts", {}).get("high", 0)),
        ("Medium Paths", summary.get("SeverityCounts", {}).get("medium", 0)),
        ("Low Paths", summary.get("SeverityCounts", {}).get("low", 0)),
        ("Informational", summary.get("SeverityCounts", {}).get("informational", 0)),
        ("MITRE Techniques", len(summary.get("MitreTechniques", []))),
        ("Attack Categories", len(summary.get("PathsByType", {}))),
        ("Evidence Sources", len(evidence)),
    ]
    for r_idx, (label, val) in enumerate(rows, 1):
        c1 = ws_sum.cell(row=r_idx, column=1, value=label)
        c2 = ws_sum.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            _apply_header(c1, header_font, header_fill)
            _apply_header(c2, header_font, header_fill)
        c1.border = thin_border
        c2.border = thin_border
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 42

    # Paths by type
    type_start = len(rows) + 2
    ws_sum.cell(row=type_start, column=1, value="Category").font = Font(bold=True)
    ws_sum.cell(row=type_start, column=2, value="Count").font = Font(bold=True)
    ws_sum.cell(row=type_start, column=3, value="Max Score").font = Font(bold=True)
    for i, (ptype, cnt) in enumerate(sorted(summary.get("PathsByType", {}).items(), key=lambda x: -x[1])):
        max_score = max((p.get("RiskScore", 0) for p in paths if p.get("Type") == ptype), default=0)
        ws_sum.cell(row=type_start + 1 + i, column=1, value=ptype)
        ws_sum.cell(row=type_start + 1 + i, column=2, value=cnt)
        ws_sum.cell(row=type_start + 1 + i, column=3, value=max_score)

    # ── Sheet 2: Attack Paths ────────────────────────────────────────
    ws_paths = wb.create_sheet("Attack Paths")
    ws_paths.sheet_properties.tabColor = "D13438"

    headers = [
        "Type", "Subtype", "Severity", "Risk Score", "Chain",
        "Chain Nodes", "Principal Name", "Principal ID",
        "Source", "Target", "Resource Name", "Resource Type",
        "Exposure", "Roles", "MITRE Technique", "MITRE Tactic",
        "Remediation", "Remediation CLI", "Remediation PowerShell",
        "Remediation Portal", "MS Learn URL",
    ]
    for c_idx, h in enumerate(headers, 1):
        _apply_header(ws_paths.cell(row=1, column=c_idx, value=h), header_font, header_fill)

    sorted_paths = sorted(paths, key=lambda p: -p.get("RiskScore", 0))
    for r_idx, p in enumerate(sorted_paths, 2):
        sev = p.get("Severity", "informational").lower()
        roles_str = ", ".join(p.get("Roles", []))

        # Format chain nodes
        nodes = p.get("ChainNodes") or []
        nodes_str = " → ".join(f'{n.get("type","")}: {n.get("label","")}' for n in nodes) if nodes else ""

        vals = [
            p.get("Type", ""),
            p.get("Subtype", ""),
            sev.upper(),
            p.get("RiskScore", 0),
            p.get("Chain", ""),
            nodes_str,
            p.get("PrincipalName", ""),
            p.get("PrincipalId", ""),
            p.get("Source", ""),
            p.get("Target", ""),
            p.get("ResourceName", ""),
            p.get("ResourceType", ""),
            p.get("Exposure", ""),
            roles_str,
            p.get("MitreTechnique", ""),
            p.get("MitreTactic", ""),
            p.get("Remediation", ""),
            p.get("RemediationCLI", ""),
            p.get("RemediationPowerShell", ""),
            p.get("RemediationPortal", ""),
            p.get("MSLearnUrl", ""),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws_paths.cell(row=r_idx, column=c_idx, value=v)
            cell.border = thin_border
        # Color the severity cell
        sev_cell = ws_paths.cell(row=r_idx, column=3)
        sev_color = _SEV_FILLS.get(sev, "FF6B6B6B")
        sev_cell.fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")
        sev_cell.font = Font(color="FFFFFF", bold=True)

    _auto_width(ws_paths, len(headers))
    ws_paths.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_paths) + 1}"
    ws_paths.freeze_panes = "A2"

    # ── Sheet 3: MITRE Coverage ──────────────────────────────────────
    ws_mitre = wb.create_sheet("MITRE Coverage")
    ws_mitre.sheet_properties.tabColor = "107C10"

    # Build real technique→tactic mapping from paths
    technique_data: dict[str, dict] = {}
    for p in paths:
        tech = p.get("MitreTechnique", "")
        tactic = p.get("MitreTactic", "")
        if not tech:
            continue
        if tech not in technique_data:
            technique_data[tech] = {"tactic": tactic, "count": 0, "max_score": 0, "severities": []}
        technique_data[tech]["count"] += 1
        technique_data[tech]["max_score"] = max(technique_data[tech]["max_score"], p.get("RiskScore", 0))
        technique_data[tech]["severities"].append(p.get("Severity", "").lower())

    mitre_headers = ["Technique", "Tactic", "Paths Count", "Max Risk Score", "Worst Severity"]
    for c_idx, h in enumerate(mitre_headers, 1):
        _apply_header(ws_mitre.cell(row=1, column=c_idx, value=h), header_font, accent_fill)

    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    for i, (tech, data) in enumerate(sorted(technique_data.items(), key=lambda x: -x[1]["max_score"]), 2):
        worst = min(data["severities"], key=lambda s: sev_order.get(s, 4))
        ws_mitre.cell(row=i, column=1, value=tech)
        ws_mitre.cell(row=i, column=2, value=data["tactic"])
        ws_mitre.cell(row=i, column=3, value=data["count"])
        ws_mitre.cell(row=i, column=4, value=data["max_score"])
        sev_cell = ws_mitre.cell(row=i, column=5, value=worst.upper())
        sc = _SEV_FILLS.get(worst, "FF6B6B6B")
        sev_cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
        sev_cell.font = Font(color="FFFFFF", bold=True)
        for c in range(1, 6):
            ws_mitre.cell(row=i, column=c).border = thin_border

    _auto_width(ws_mitre, len(mitre_headers))
    ws_mitre.freeze_panes = "A2"

    # ── Sheet 4: Remediation Tracker ─────────────────────────────────
    ws_rem = wb.create_sheet("Remediation Tracker")
    ws_rem.sheet_properties.tabColor = "FF8C00"

    rem_headers = [
        "Priority", "Severity", "Attack Path Type", "Subtype",
        "Remediation Action", "CLI Command", "PowerShell Command",
        "MS Learn Link", "Status", "Owner", "Due Date", "Notes",
    ]
    for c_idx, h in enumerate(rem_headers, 1):
        _apply_header(
            ws_rem.cell(row=1, column=c_idx, value=h),
            header_font,
            PatternFill(start_color="FFFF8C00", end_color="FFFF8C00", fill_type="solid"),
        )

    # Deduplicate remediation actions by (type, subtype, remediation)
    seen_rem: set[tuple[str, str, str]] = set()
    rem_rows: list[dict] = []
    for p in sorted_paths:
        key = (p.get("Type", ""), p.get("Subtype", ""), p.get("Remediation", ""))
        if key in seen_rem or not key[2]:
            continue
        seen_rem.add(key)
        rem_rows.append(p)

    for r_idx, p in enumerate(rem_rows, 2):
        sev = p.get("Severity", "informational").lower()
        vals = [
            r_idx - 1,  # Priority
            sev.upper(),
            p.get("Type", ""),
            p.get("Subtype", ""),
            p.get("Remediation", ""),
            p.get("RemediationCLI", ""),
            p.get("RemediationPowerShell", ""),
            p.get("MSLearnUrl", ""),
            "Not Started",  # Status placeholder
            "",  # Owner placeholder
            "",  # Due Date placeholder
            "",  # Notes placeholder
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws_rem.cell(row=r_idx, column=c_idx, value=v)
            cell.border = thin_border
        # Severity color
        sev_cell = ws_rem.cell(row=r_idx, column=2)
        sc = _SEV_FILLS.get(sev, "FF6B6B6B")
        sev_cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
        sev_cell.font = Font(color="FFFFFF", bold=True)
        # Status cell with data validation style
        status_cell = ws_rem.cell(row=r_idx, column=9)
        status_cell.fill = PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid")

    _auto_width(ws_rem, len(rem_headers))
    ws_rem.auto_filter.ref = f"A1:{get_column_letter(len(rem_headers))}{len(rem_rows) + 1}"
    ws_rem.freeze_panes = "A2"

    # ── Sheet 5: Compliance Coverage ─────────────────────────────────
    ws_comp = wb.create_sheet("Compliance Coverage")
    ws_comp.sheet_properties.tabColor = "5C2D91"

    comp_headers = ["Attack Path Type", "Subtype", "Severity", "Risk Score"] + list(_FRAMEWORKS) + ["Frameworks Hit"]
    comp_fill = PatternFill(start_color="FF5C2D91", end_color="FF5C2D91", fill_type="solid")
    for c_idx, h in enumerate(comp_headers, 1):
        _apply_header(ws_comp.cell(row=1, column=c_idx, value=h), header_font, comp_fill)

    for r_idx, p in enumerate(sorted_paths, 2):
        sev = p.get("Severity", "informational").lower()
        fw = p.get("ComplianceFrameworks") or {}
        fw_count = sum(1 for f in _FRAMEWORKS if fw.get(f))
        vals = [
            p.get("Type", ""),
            p.get("Subtype", ""),
            sev.upper(),
            p.get("RiskScore", 0),
        ]
        for f in _FRAMEWORKS:
            controls = fw.get(f, [])
            vals.append(", ".join(controls) if controls else "—")
        vals.append(fw_count)

        for c_idx, v in enumerate(vals, 1):
            cell = ws_comp.cell(row=r_idx, column=c_idx, value=v)
            cell.border = thin_border
        # Severity color
        sev_cell = ws_comp.cell(row=r_idx, column=3)
        sc = _SEV_FILLS.get(sev, "FF6B6B6B")
        sev_cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
        sev_cell.font = Font(color="FFFFFF", bold=True)
        # Color framework cells
        for fi, f in enumerate(_FRAMEWORKS, 5):
            cell = ws_comp.cell(row=r_idx, column=fi)
            if fw.get(f):
                cell.fill = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")

    _auto_width(ws_comp, len(comp_headers))
    ws_comp.auto_filter.ref = f"A1:{get_column_letter(len(comp_headers))}{len(sorted_paths) + 1}"
    ws_comp.freeze_panes = "A2"

    # ── Sheet 6: Evidence Summary ────────────────────────────────────
    ws_ev = wb.create_sheet("Evidence Summary")
    ws_ev.sheet_properties.tabColor = "008272"

    ev_headers = ["Evidence Source", "Item Count", "Status", "Sample Keys"]
    ev_fill = PatternFill(start_color="FF008272", end_color="FF008272", fill_type="solid")
    for c_idx, h in enumerate(ev_headers, 1):
        _apply_header(ws_ev.cell(row=1, column=c_idx, value=h), header_font, ev_fill)

    for r_idx, (src, data) in enumerate(sorted(evidence.items()), 2):
        items = data if isinstance(data, list) else [data] if data else []
        sample_keys = ""
        if items and isinstance(items[0], dict):
            sample_keys = ", ".join(list(items[0].keys())[:6])
        ws_ev.cell(row=r_idx, column=1, value=src).border = thin_border
        ws_ev.cell(row=r_idx, column=2, value=len(items)).border = thin_border
        status = "✓ Collected" if items else "⚠ Empty"
        status_cell = ws_ev.cell(row=r_idx, column=3, value=status)
        status_cell.border = thin_border
        if items:
            status_cell.fill = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid")
        ws_ev.cell(row=r_idx, column=4, value=sample_keys).border = thin_border

    _auto_width(ws_ev, len(ev_headers))

    wb.save(output_path)
