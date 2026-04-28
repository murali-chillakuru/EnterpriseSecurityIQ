"""Excel export — write findings + evidence to .xlsx with openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)


def export_excel(
    findings: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """Write findings + evidence to a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # -- Findings sheet --
    ws = wb.active
    ws.title = "Findings"
    finding_cols = [
        "FindingId", "ControlId", "Framework", "ControlTitle",
        "Status", "Severity", "Domain", "Description", "Recommendation",
    ]
    _write_header(ws, finding_cols)
    for row_idx, f in enumerate(findings, start=2):
        for col_idx, col in enumerate(finding_cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(f.get(col, "")))

    # -- Evidence sheet --
    ws2 = wb.create_sheet("Evidence")
    evidence_cols = [
        "EvidenceId", "EvidenceType", "Source", "Collector",
        "Description", "CollectedAt", "ResourceId",
    ]
    _write_header(ws2, evidence_cols)
    for row_idx, e in enumerate(evidence, start=2):
        for col_idx, col in enumerate(evidence_cols, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=str(e.get(col, "")))

    wb.save(str(output_path))
    log.info("Excel exported: %s", output_path)
    return output_path


def _write_header(ws, columns: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(name) + 4, 15)
